import os
from flask import Flask, render_template, request, redirect, url_for, jsonify
import openpyxl
import pandas as pd

app = Flask(__name__, template_folder='.')

UPLOAD_FOLDER = 'uploads'
EXCEL_FILE = 'etudiants.xlsx'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Route pour afficher la page d'accueil (index.html)
@app.route('/')
def home():
    return render_template('index.html')

# Route pour afficher la page d'ajout (ajout.html)
@app.route('/add_etudiant')
def add_etudiant_form():
    return render_template('ajout.html')

@app.route('/add_liste_etudiant')
def add_liste_etudiant_form():
    # return render_template('ajout_file.html')
    return render_template('ajout_file.html')

@app.route('/get_complete_upload')
def get_complete_upload():
    csv_path = request.args.get('csv_path')
    csv_data = pd.read_csv(csv_path)
    # print("list: ---------------")
    # print(csv_data.values.tolist())
    # print("---------------------")

    return render_template('confirm_file.html', students=csv_data.values.tolist(), csv_path=csv_path)

# Route pour afficher la page "Our Team" (ourteam.html)
@app.route('/ourteam')
def our_team():
    return render_template('ourteam.html')

@app.route('/list_etudiants') 
def list_etudiants(): 
    workbook = openpyxl.load_workbook('etudiants.xlsx') 
    sheet = workbook.active 
    students = [] 
    for row in sheet.iter_rows(min_row=2, values_only=True):    
         students.append(row) 
    return render_template('tables.html',students=students)

# Route pour ajouter un étudiant depuis le formulaire dans ajout.html
@app.route('/add_etudiant', methods=['POST'])
def add_etudiant():
    if request.is_json:
        data = request.get_json()
        nom = data['nom']
        prenom = data['prenom']
        age = data['age']
        email = data['email']
        gouvernorat = data['gouvernorat']
        specialite = data['specialite']
        moyenne = data['moyenne']
        

        # Ouvrir ou créer le fichier Excel
        try:
            workbook = openpyxl.load_workbook('etudiants.xlsx')
            sheet = workbook.active
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Nom", "Prenom", "Age", "Email", "Gouvernorat", "Specialité", "Moyenne"])  # En-têtes

        # Ajouter les données dans le fichier Excel
        sheet.append([nom, prenom, age, email, gouvernorat, specialite, moyenne])
        workbook.save('etudiants.xlsx')

        gouvernorat = data['gouvernorat']
        specialite = data['specialite']
        moyenne = data['moyenne']
        return jsonify({"message": "Étudiant ajouté avec succès."}), 200
    else:
        return jsonify({"message": "Format de données non supporté."}), 400
    


@app.route('/add_liste_etudiant', methods=['POST'])
def add_liste_etudiant():
    # Check if a file is in the request
    if 'csv_file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400

    file = request.files['csv_file']

    # Check if a file was actually selected
    if file.filename == '':
        return jsonify({"error": "No selected file"}), 400

    # Ensure it's a CSV file
    if not file.filename.endswith('.csv'):
        return jsonify({"error": "File type not allowed"}), 400

    # Save the CSV temporarily
    csv_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(csv_path)

    #return jsonify({"message": "CSV data appended to Excel file"}), 200
    return redirect(url_for('get_complete_upload', csv_path=csv_path))

@app.route('/complete_upload', methods=['POST'])
def complete_upload():
    csv_path = request.form.get('csv_path')
    concat = request.form.get('concat')
    # Handle the uploaded file (e.g., process CSV)

    if concat:
        # Read the CSV file
        try:
            csv_data = pd.read_csv(csv_path)
        except Exception as e:
            return jsonify({"error": f"Failed to read CSV file: {str(e)}"}), 500

        # Open or create the Excel file
        if os.path.exists(EXCEL_FILE):
            # Load existing workbook and append data
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                for col in csv_data.columns:
                    csv_data[col] = csv_data[col].astype(str)  # Convert to string to avoid mixed type errors
                csv_data.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        else:
            # Create new workbook and write data with headers
            with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
                csv_data.to_excel(writer, index=False, header=True)

    # Remove the temporary CSV file
    os.remove(csv_path)

    # return jsonify({"message": "File uploaded successfully", "csv_path": csv_path}), 200
    return redirect(url_for('list_etudiants', csv_path=csv_path))

if __name__ == '__main__':
    app.run(debug=True)
